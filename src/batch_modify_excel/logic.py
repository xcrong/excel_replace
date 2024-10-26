import os
import json
import threading
from tkinter import messagebox
from multiprocessing import Pool, cpu_count

from openpyxl import load_workbook
from openpyxl.utils import (
    get_column_letter,
    column_index_from_string,
)

from batch_modify_excel.config import Config, logger


class ExcelReplicerLogic:
    def __init__(self):
        self.config_file = Config.CONFIG_FILE
        self.config = self.load_config()
        self.processing = False
        self.current_thread = None
        self.process_pool_size = max(1, cpu_count() - 1)

    def load_config(self):
        """从文件中加载配置"""
        if os.path.exists(self.config_file):
            with open(self.config_file, "r", encoding="utf-8") as f:
                return json.load(f)
        return {}

    def save_config(self, values):
        """将配置保存到文件中"""
        try:
            self.config.update(values)
            with open(self.config_file, "w", encoding="utf-8") as f:
                json.dump(
                    self.config,
                    f,
                    ensure_ascii=False,
                    indent=4,
                )
        except Exception as e:
            logger.info(Config.CONFIG_SAVE_ERROR.format(str(e)))

    def parse_input(self, input_str: str, is_row=False):
        """将输入字符串解析为列表"""
        input_str = input_str.strip()
        if not input_str:
            return None

        for sep in Config.SEPARATORS:
            if sep in input_str:
                items = [item.strip() for item in input_str.split(sep)]
                if is_row:
                    try:
                        return [int(item) for item in items]
                    except ValueError:
                        messagebox.showerror(
                            Config.INPUT_ERROR_MESSAGE,
                            Config.ROW_NUMBER_ERROR,
                        )
                        return None
                return items
        return [input_str.strip()]

    @staticmethod
    def validate_column(col):
        """验证并转换列标识"""
        try:
            # 如果是数字，转换为字母
            if col.isdigit():
                return get_column_letter(int(col))
            # 确保列标识是有效的
            column_index_from_string(col.upper())
            return col.upper()
        except ValueError:
            return None

    @staticmethod
    def process_cell_value(cell_value, target_char, replace_char):
        """处理单元格值"""
        if cell_value is None:
            return None
        return str(cell_value).replace(target_char, replace_char)

    @staticmethod
    def process_sheet(sheet, columns, rows, target_char, replace_char):
        """处理单个工作表"""
        # 获取该工作表的实际维度
        max_row = sheet.max_row
        max_column = sheet.max_column

        # 处理指定列
        if columns:
            logger.info(f"指定列{columns}")
            valid_columns = []
            for col in columns:
                valid_col = ExcelReplicerLogic.validate_column(col)
                if valid_col:
                    col_idx = column_index_from_string(valid_col)
                    if col_idx <= max_column:  # 确保列在工作表范围内
                        valid_columns.append(valid_col)

            for col in valid_columns:
                for row in range(1, max_row + 1):
                    cell = sheet[f"{col}{row}"]
                    if cell.value and target_char in str(cell.value):
                        cell.value = ExcelReplicerLogic.process_cell_value(
                            cell.value,
                            target_char,
                            replace_char,
                        )

        # 处理指定行
        if rows:
            logger.info(f"指定行{rows}")
            valid_rows = [r for r in rows if 1 <= r <= max_row]
            for row in valid_rows:
                for col_idx in range(1, max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    cell = sheet[f"{col_letter}{row}"]
                    if cell.value and target_char in str(cell.value):
                        cell.value = ExcelReplicerLogic.process_cell_value(
                            cell.value,
                            target_char,
                            replace_char,
                        )

        # 如果没有指定行和列，处理所有单元格
        if not columns and not rows:
            logger.info("处理所有单元格")
            for row in range(1, max_row + 1):
                for col_idx in range(1, max_column + 1):
                    col_letter = get_column_letter(col_idx)
                    cell = sheet[f"{col_letter}{row}"]
                    if cell.value and target_char in str(cell.value):
                        cell.value = ExcelReplicerLogic.process_cell_value(
                            cell.value,
                            target_char,
                            replace_char,
                        )

    @staticmethod
    def replace_target_char_worker(args):
        """
        工作进程的处理函数
        返回元组 (是否成功, 文件路径)
        """
        (
            file_path,
            columns,
            rows,
            target_char,
            replace_char,
            overwrite,
        ) = args
        try:
            wb = load_workbook(filename=file_path)

            # 遍历每个工作表
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                logger.info(f"处理工作表 {sheet_name}")

                # 处理当前工作表
                ExcelReplicerLogic.process_sheet(
                    sheet,
                    columns,
                    rows,
                    target_char,
                    replace_char,
                )

            # 保存文件
            save_path = (
                file_path
                if overwrite
                else os.path.join(
                    os.path.dirname(file_path),
                    "new_" + os.path.basename(file_path),
                )
            )

            wb.save(save_path)
            wb.close()  # 确保关闭工作簿
            logger.info(Config.FILE_SAVED_MESSAGE.format(save_path))
            return True, file_path
        except Exception as e:
            logger.error(f"处理文件 {file_path} 时出错: {str(e)}")
            return False, file_path

    def process_files(self, values, progress_callback, complete_callback):
        """使用进程池处理文件"""
        file_path = values["last_file_path"]
        columns = self.parse_input(values["columns"])
        rows = self.parse_input(values["rows"], is_row=True)
        target_char = values["target_char"]
        replace_char = values["replace_char"]
        overwrite = values["overwrite"]

        try:
            if values["select_folder"]:
                # 收集所有需要处理的文件
                excel_files = []
                for root, _, files in os.walk(file_path):
                    excel_files.extend(
                        [os.path.join(root, f) for f in files if f.endswith(".xlsx")]
                    )

                total_files = len(excel_files)
                if total_files == 0:
                    complete_callback(0, 0)
                    return

                # 准备进程池的参数
                process_args = [
                    (
                        f,
                        columns,
                        rows,
                        target_char,
                        replace_char,
                        overwrite,
                    )
                    for f in excel_files
                ]

                processed_files = 0
                success_count = 0

                # 使用进程池处理文件
                with Pool(processes=self.process_pool_size) as pool:
                    for (
                        success,
                        file_path,
                    ) in pool.imap_unordered(
                        self.replace_target_char_worker,
                        process_args,
                    ):
                        if not self.processing:
                            pool.terminate()
                            break

                        if success:
                            success_count += 1
                        processed_files += 1
                        progress_callback(
                            processed_files,
                            total_files,
                            os.path.basename(file_path),
                        )

                complete_callback(success_count, processed_files)
            else:
                # 单文件处理
                filename = os.path.basename(file_path)
                success, _ = self.replace_target_char_worker(
                    (
                        file_path,
                        columns,
                        rows,
                        target_char,
                        replace_char,
                        overwrite,
                    )
                )
                progress_callback(1, 1, filename)
                complete_callback(1 if success else 0, 1)

        except Exception as e:
            logger.error(f"处理文件时出错: {str(e)}")
            complete_callback(0, 1)
        finally:
            self.processing = False

    def execute_cleaning(self, values, progress_callback, complete_callback):
        """在单独的线程中启动多进程处理"""
        if self.processing:
            return False

        self.processing = True
        self.save_config(values)

        self.current_thread = threading.Thread(
            target=self.process_files,
            args=(
                values,
                progress_callback,
                complete_callback,
            ),
        )
        self.current_thread.daemon = True
        self.current_thread.start()
        return True

    def stop_processing(self):
        """停止当前的处理过程"""
        self.processing = False
        if self.current_thread and self.current_thread.is_alive():
            self.current_thread.join(timeout=1.0)
